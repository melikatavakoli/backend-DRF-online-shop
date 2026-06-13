from __future__ import annotations
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from users.serializers import StudentListSerializer
from users.choices import FieldChoices, GradeChoices


def build_student_export_payload(queryset):
    serialized_data = StudentListSerializer(queryset, many=True).data

    columns = [
        ("id", "شناسه"),
        ("full_name", "نام و نام خانوادگی"),
        ("mobile", "شماره موبایل"),
        ("grade", "پایه"),
        ("field", "رشته"),
        ("description", "توضیحات"),
        ("_created_at", "تاریخ ثبت"),
    ]

    rows = []
    for item in serialized_data:
        grade_value = item.get("grade")
        field_value = item.get("field")

        try:
            grade_label = (
                GradeChoices(grade_value).label if grade_value is not None else "-"
            )
        except ValueError:
            grade_label = grade_value

        try:
            field_label = FieldChoices(field_value).label if field_value else "-"
        except ValueError:
            field_label = field_value

        rows.append(
            {
                "id": item.get("id"),
                "full_name": item.get("full_name", "-"),
                "mobile": item.get("mobile", "-"),
                "grade": grade_label,
                "field": field_label,
                "description": item.get("description", "-"),
            }
        )

    return {
        "kind": "excel",
        "filename": "students-report.xlsx",
        "sheet_title": "لیست دانش آموزان",
        "columns": columns,
        "rows": rows,
    }


def export_students_to_excel(queryset):
    payload = build_student_export_payload(queryset)

    columns = payload["columns"]
    rows = payload["rows"]
    sheet_title = payload["sheet_title"]

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_title
    worksheet.sheet_view.rightToLeft = True

    header_font = Font(name="Calibri", size=11, bold=True)
    body_font = Font(name="Calibri", size=10)
    header_alignment = Alignment(horizontal="center", vertical="center")
    body_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    col_widths = []

    for col_idx, (_, title) in enumerate(columns, start=1):
        cell = worksheet.cell(row=1, column=col_idx, value=title)
        cell.font = header_font
        cell.alignment = header_alignment
        col_widths.append(len(str(title)))

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, (key, _) in enumerate(columns, start=1):
            value = row.get(key, "")

            cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
            cell.font = body_font
            cell.alignment = body_alignment

            display_len = len(str(value or ""))
            if display_len > col_widths[col_idx - 1]:
                col_widths[col_idx - 1] = display_len

    for idx, width in enumerate(col_widths, start=1):
        worksheet.column_dimensions[get_column_letter(idx)].width = min(
            max(width + 2, 12), 80
        )

    worksheet.freeze_panes = "A2"

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    return output
